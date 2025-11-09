from openpyxl import load_workbook

def read_xlsx(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if any(cell is None for cell in row):
            continue
        rows.append(row)
    return rows

if __name__ == "__main__":
    read_xlsx(r"Excel Files/Basvurular.xlsx")